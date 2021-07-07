from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import xlsxwriter
import logging

#Solo tocar esto:
path = r"C:\Users\janic\OneDrive\Documentos\Meli\dat.xlsx"

class URLAlreadyOnExcel(Exception):
    def __init__(self, message):
        self.__message = message

def url_already_on_excel(url, sheet):
    for row in range(1,sheet.max_row+1):
        if sheet.cell(row, 1).value == url:
            return True
    return False

def put_excel_first_line(Sheet1, soup):
    first_row = ["URL", "Title", "Job Description", "Skills and Requirements", "Pluses"]
    if Sheet1.max_row == 1:
        add_new_row(first_row, 1, Sheet1)

def add_new_row(my_list, my_row, Sheet1):
    my_column = 1
    for my_element in my_list:
        Sheet1.cell(row = my_row, column = my_column).value = my_element
        my_column = my_column + 1

def save_row(my_row_list, soup):
    wb = load_workbook(path)
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]
    put_excel_first_line(Sheet1, soup)
    if url_already_on_excel(my_row_list[0], Sheet1):
        raise URLAlreadyOnExcel("Este URL ya se encuentra en el excel.")
    first_empty_row = Sheet1.max_row + 1
    add_new_row(my_row_list, first_empty_row, Sheet1)
    wb.save(path)
    print("Ya se guardo la data del URL: " + my_row_list[0])

def get_skills_and_requirements(soup):
    number = 0
    skills_and_requirements= []
    for row_element in soup.find_all('li'):
        if number >= 64 and row_element.text != "GalleryCommunityTrending" :
            skills_and_requirements.append(row_element.text)
        if row_element.text == "GalleryCommunityTrending":
            break
        number = number + 1
    return skills_and_requirements

def list_to_string(list):
    my_str = ""
    for elem in list:
        my_str = my_str + elem
        my_str = my_str + " \n"
    return my_str

def get_pluses(skills_and_requirements):
    skills = list_to_string(skills_and_requirements)
    if "Pluses: " in skills:
        return skills.split("Pluses: ")
    return [skills]

def get_row_from_page(soup, url):
    my_row = []
    my_row.append(url)
    my_row.append(soup.find_all('title')[0].text)
    my_row.append(soup.find_all('p')[2].text)
    skills = get_pluses(get_skills_and_requirements(soup))
    my_row.append(skills[0])
    if len(skills)==2:
        my_row.append(skills[1])
    return my_row

while(1):
    url = input("Ingresa el url: ").strip()
    if url == "exit":
        exit()
    try:
        html_content = requests.get(url).text
        soup = BeautifulSoup(html_content, "lxml")
        save_row(get_row_from_page(soup, url), soup)
    except PermissionError:
        print("ERROR: Tenes que cerrar el archivo excel para poder ejecutar el programa.")
    except requests.exceptions.MissingSchema:
        print("ERROR: El URL no existe o no se puede acceder a el.")
    except URLAlreadyOnExcel as e:
        print("ERROR: El URL ya se encuentra en el excel")
